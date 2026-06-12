#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# 把 data/profit_mining/每日自选股清单.csv 渲染成 Excel(.xlsx) 工作簿。
#   单工作表"每日选股":顶部概览摘要 + 表头 + 三档(核心/精选/其余)顺序 + 精选行底色高亮,
#   冻结表头、AutoFilter、列宽自适应、代码列保前导0(文本)。
#   数据读取复用 export_watchlist_md.read_rows(同源、同 zfill)。需要 openpyxl。
# 用法: export_watchlist_xlsx.py [--csv <路径>] [--out <xlsx路径>]   省略 --out 默认写到 CSV 同目录。
import sys, io, os, re, glob, datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import export_watchlist_md as EXP   # 复用 read_rows / 同源 SHOW 思路

DEFAULT_CSV = EXP.DEFAULT_CSV

# (CSV列名, 表头) —— 与 Markdown 文档同列(含板块)
SHOW = [("精选", "精选"), ("星级", "星级"), ("可入状态", "可入"), ("优先级", "优先级"), ("命中规则", "规则"),
        ("股票代码", "代码"), ("股票名称", "名称"), ("板块", "板块"),
        ("买点类型", "买点"), ("信号日期", "信号日期"),
        ("买入价", "买入价"), ("扫描日价", "扫描价"), ("止损价", "止损价"), ("止盈价", "止盈价"),
        ("预估胜率", "胜率%"), ("量比", "量比"), ("资金确认", "资金"), ("中枢底部", "中枢底"),
        ("获利盘%", "获利盘%"), ("大盘环境", "大盘")]

# 各列宽度(字符)
WIDTHS = {"精选": 8, "优先级": 9, "命中规则": 9, "股票代码": 9, "股票名称": 11,
          "板块": 10, "买点类型": 7, "信号日期": 12, "买入价": 9, "止损价": 9,
          "止盈价": 9, "量比": 7, "资金确认": 8, "中枢底部": 8, "获利盘%": 9, "大盘环境": 8}

TEXT_COLS = {"股票代码"}   # 强制文本以保前导0


def fill_sheet(ws, rows, slot_label=""):
    """把一段清单写入给定 worksheet ws。slot_label 非空时概览行带"时段 X"。"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    nA = sum(1 for r in rows if "A" in r.get("命中规则", ""))
    nB = sum(1 for r in rows if "B" in r.get("命中规则", ""))
    nCore = sum(1 for r in rows if r.get("精选") == "★★核心")
    nSel = sum(1 for r in rows if r.get("精选"))
    nZ = sum(1 for r in rows if r.get("资金确认"))
    nZS = sum(1 for r in rows if r.get("中枢底部"))
    scan = rows[0].get("扫描日期", "") if rows else ""
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

    core = [r for r in rows if r.get("精选") == "★★核心"]
    sel = [r for r in rows if r.get("精选") == "★精选"]
    rest = [r for r in rows if not r.get("精选")]
    ordered = core + sel + rest
    ncol = len(SHOW)

    head_fill = PatternFill("solid", fgColor="C9A84A")
    head_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    core_fill = PatternFill("solid", fgColor="FFE9A8")
    sel_fill = PatternFill("solid", fgColor="FFF6D5")
    star_font = Font(name="微软雅黑", bold=True, color="B8860B")
    name_font = Font(name="微软雅黑", bold=True)
    base_font = Font(name="微软雅黑", size=10)
    loss_font = Font(name="微软雅黑", size=10, bold=True, color="C0392B")
    gain_font = Font(name="微软雅黑", size=10, bold=True, color="1E7E34")
    thin = Side(style="thin", color="E0D8C0")
    border = Border(bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    slot_prefix = f"时段 {slot_label}　" if slot_label else ""
    summary = (f"🛡️ 每日稳定选股　{slot_prefix}扫描日期 {scan}　共 {len(rows)} 只"
               f"（A抄底 {nA} / B抢筹 {nB}）　精选 {nSel}（★★核心 {nCore}）"
               f"　资金确认 {nZ}　中枢底部 {nZS}　|　生成 {now}")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    c = ws.cell(row=1, column=1, value=summary)
    c.font = Font(name="微软雅黑", bold=True, color="4A3A1A", size=11)
    c.alignment = left
    c.fill = PatternFill("solid", fgColor="F6F1E4")
    ws.row_dimensions[1].height = 22

    for j, (_, title) in enumerate(SHOW, start=1):
        cell = ws.cell(row=2, column=j, value=title)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = center

    for i, r in enumerate(ordered, start=3):
        tier = r.get("精选", "")
        fill = core_fill if tier == "★★核心" else (sel_fill if tier == "★精选" else None)
        for j, (col, _) in enumerate(SHOW, start=1):
            val = r.get(col, "")
            cell = ws.cell(row=i, column=j)
            if col in TEXT_COLS:
                cell.value = str(val)
                cell.number_format = "@"
            else:
                cell.value = val if val not in (None, "") else ""
            cell.border = border
            cell.alignment = left if col in ("股票名称", "板块") else center
            if fill:
                cell.fill = fill
            if col == "精选" and tier:
                cell.font = star_font
            elif col == "股票名称":
                cell.font = name_font
            elif col == "止损价":
                cell.font = loss_font
            elif col == "止盈价":
                cell.font = gain_font
            else:
                cell.font = base_font

    for j, (col, _) in enumerate(SHOW, start=1):
        ws.column_dimensions[get_column_letter(j)].width = WIDTHS.get(col, 10)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ncol)}{2 + len(ordered)}"
    return ws


def build_workbook(rows, slot_label=""):
    """单 sheet 工作簿（向后兼容）。"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "每日选股"
    fill_sheet(ws, rows, slot_label)
    return wb


def build_multi_sheet_workbook(slots):
    """slots=[(sheet名, rows), …] → 多 sheet 工作簿（删默认空 sheet）。"""
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in slots:
        ws = wb.create_sheet(title=name[:31])
        fill_sheet(ws, rows, name)
    return wb


def resolve_data_paths(csv_path):
    """从传入 csv 路径推导 (盘中时段目录, 盘后latest路径)，容器/宿主通用。"""
    p = os.path.abspath(csv_path)
    parts = p.split(os.sep)
    if "profit_mining" in parts:
        pm = os.sep.join(parts[:parts.index("profit_mining") + 1])
        return (os.path.join(pm, "watchlist_history", "intraday"),
                os.path.join(pm, "每日自选股清单.csv"))
    return (os.path.dirname(p), p)   # 兜底：未知结构（如测试自定义路径）


def collect_today_slots(scan_date, include_eod, intraday_dir, eod_csv):
    """汇集当天各时段(+可选盘后)清单 → [(sheet名, rows), …]，按时间升序。"""
    items = []
    pattern = os.path.join(intraday_dir, f"每日自选股清单_{scan_date}_*.csv")
    for path in glob.glob(pattern):
        m = re.search(r"_(\d{4})\.csv$", os.path.basename(path))
        if not m:
            continue
        hhmm = m.group(1)
        label = f"{hhmm[:2]}：{hhmm[2:]}"   # 全角冒号：Excel sheet 名禁用半角:
        items.append((hhmm, label, path))
    slots = []
    for _, label, path in sorted(items):
        rows = EXP.read_rows(path)
        if rows:
            slots.append((label, rows))
    if include_eod and os.path.exists(eod_csv):
        eod_rows = EXP.read_rows(eod_csv)
        if eod_rows and eod_rows[0].get("扫描日期") == scan_date:
            slots.append(("盘后", eod_rows))
    return slots


def build_xlsx_bytes(rows):
    """返回 .xlsx 的二进制内容(供邮件附件)。"""
    wb = build_workbook(rows)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_xlsx_bytes_multi(slots):
    """多 sheet 工作簿 → .xlsx 二进制（供邮件附件）。"""
    wb = build_multi_sheet_workbook(slots)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def main():
    args = sys.argv[1:]
    csv_path, out = DEFAULT_CSV, None
    i = 0
    while i < len(args):
        if args[i] == "--csv":
            csv_path = args[i + 1]; i += 2
        elif args[i] == "--out":
            out = args[i + 1]; i += 2
        else:
            print(f"未知参数 {args[i]}", file=sys.stderr); sys.exit(2)
    rows = EXP.read_rows(csv_path)
    if not rows:
        print("清单为空", file=sys.stderr); sys.exit(1)
    scan = rows[0].get("扫描日期", "")
    intraday_dir, eod_csv = resolve_data_paths(csv_path)
    slots = collect_today_slots(scan, include_eod=True,
                                intraday_dir=intraday_dir, eod_csv=eod_csv)
    if not slots:
        slots = [("盘后", rows)]   # 兜底：当天无任何时段 CSV
    if out is None:
        out = os.path.join(os.path.dirname(csv_path), f"每日稳定选股_{scan}.xlsx")
    wb = build_multi_sheet_workbook(slots)
    wb.save(out)
    nSel = sum(1 for r in rows if r.get("精选"))
    nCore = sum(1 for r in rows if r.get("精选") == "★★核心")
    print(f"已写 Excel 文档 -> {out}（{scan} 共{len(rows)}只/精选{nSel}/核心{nCore}）")


if __name__ == "__main__":
    main()
